# GRAPHICS LIBRARY
from graphics import *
# ASSIGN OBJECT
from copy import deepcopy
# MATH LIBRARY
import math
# EXCEL LEARN LIBRARY
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
# DEBOUNCE TIME
import time


# ------------------------- PRIMITIVE GRAPHICS CLASS -------------------------  #
class PrimitiveGraphs:
    
    # --- CONSTRUCTOR CLASS --- #
    def __init__(self, width = 800, height = 600):
        # STARTER GRAPHICS LIBRARY
        self.win = GraphWin("Primitive Graphics", width, height)
        self.win.setBackground(color_rgb(16,16,37))

        # DEFINE MAX POSITION IN GRAPHS
        self.x_max = width
        self.y_max = height

        # ZBUFFER ARRAY STARTER POSITIONS
        self.z_buffer = []
        for i in range(width):
            self.z_buffer.append(["empty" for j in range(height)])

        # STARTER EXCEL LEARN LIBRARY
        wb = load_workbook(filename = 'plnilha de radar.xlsx', data_only=True)
        sheet = wb.active

        max_row = sheet.max_row
        max_col = sheet.max_column

        # LEARN EXCEL AND SAVE CELL VALUES
        self.datas = []

        for row in range(2, max_row + 1):
            data_row = []

            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)

                if(cell.value == None):
                    continue

                data_row.append(cell.value)
            
            if(len(data_row) == 0):
                continue

            self.datas.append(data_row)


    # --- RENDER POINT METHOD --- #
    def point(self, x, y, color, width):
        x = x + self.x_max/ 2
        y = self.y_max / 2 - y

        # POINT WIDTH
        if width == 1:
            self.win.plotPixel(x, y, color)

        if width == 2:
            self.win.plotPixel(x, y, color)
            self.win.plotPixel(x+1, y, color)
            self.win.plotPixel(x, y-1, color)
            self.win.plotPixel(x+1, y-1, color)

        if width == 3:
            self.win.plotPixel(x, y, color)
            self.win.plotPixel(x+1, y, color)
            self.win.plotPixel(x, y-1, color)
            self.win.plotPixel(x+1, y-1, color)
            self.win.plotPixel(x+2, y, color)
            self.win.plotPixel(x, y-2, color)
            self.win.plotPixel(x+2, y-2, color)
        
        if((x < self.x_max and x > 0) and (y < self.y_max and y > 0)):
            self.z_buffer[int(x)][int(y)] = color
        else:
            print(x, y)


    def accuratePoint(self, x, y, color, width):

        x = x + self.x_max/ 2
        y = self.y_max / 2 - y
        
        for increment in range(width):
            self.win.plotPixel(x, y, color)

            self.win.plotPixel(x+increment, y, color)
            self.win.plotPixel(x-increment, y, color)

            self.win.plotPixel(x, y+increment, color)
            self.win.plotPixel(x, y-increment, color)

            self.win.plotPixel(x+increment, y-increment, color)
            self.win.plotPixel(x-increment, y+increment, color)

            self.win.plotPixel(x+increment, y+increment, color)
            self.win.plotPixel(x-increment, y-increment, color)

            for sub_increment in range(increment):
                self.win.plotPixel(x+increment, y+sub_increment, color)
                self.win.plotPixel(x-increment, y-sub_increment, color)

                self.win.plotPixel(x-increment, y+sub_increment, color)
                self.win.plotPixel(x+increment, y-sub_increment, color)

                self.win.plotPixel(x+sub_increment, y+increment, color)
                self.win.plotPixel(x-sub_increment, y-increment, color)

                self.win.plotPixel(x-sub_increment, y+increment, color)
                self.win.plotPixel(x+sub_increment, y-increment, color)

                sub_increment += 1

            increment += 1

        try:
            self.z_buffer[int(x)][int(y)] = color
        finally:
            pass


    def straight(self, x1, y1, x2, y2, color, width, style = "continuos"):
        x1 = int(x1)
        y1 = int(y1)
        x2 = int(x2)
        y2 = int(y2)

        x = x1
        y = y1
        p = 0
        d_x = x2 - x1
        d_y = y2 - y1
        count = 0
        lig = True

        x_inc = 1
        y_inc = 1

        if d_x < 0:
            x_inc = - x_inc
            d_x = - d_x

        if d_y < 0:
            y_inc = - y_inc
            d_y = - d_y

        if d_y <= d_x:
            p = d_x / 2

            while x != x2:
                if(style == "dotted"):
                    if(count == 5):
                        count = 0
                        lig = not lig

                    if(lig):
                        self.point(x, y, color, width)
                        count += 1
                        
                    else:
                        count += 1
                        if(count == 5):
                            lig = not lig
                            count = 0

                elif(style == "dashed"):
                    if(count == 15):
                        count = 0
                        lig = not lig

                    if(lig):
                        self.point(x, y, color, width)
                        count += 1
                        
                    else:
                        count += 1
                        if(count == 5):
                            lig = not lig
                            count = 0
                    
                else:
                    self.point(x, y, color, width)

                p = p - d_y

                if p < 0:
                    y = y + y_inc
                    p = p + d_x

                x = x + x_inc


        else:
            p = d_y / 2

            while y != y2:
                if(style == "dotted"):
                    if(count == 5):
                        count = 0
                        lig = not lig

                    if(lig):
                        self.point(x, y, color, width)
                        count += 1
                        
                    else:
                        count += 1
                        if(count == 5):
                            lig = not lig
                            count = 0

                elif(style == "dashed"):
                    if(count == 15):
                        count = 0
                        lig = not lig

                    if(lig):
                        self.point(x, y, color, width)
                        count += 1
                        
                    else:
                        count += 1
                        if(count == 5):
                            lig = not lig
                            count = 0

                else:
                    self.point(x, y, color, width)

                p = p - d_x

                if p < 0:
                    x = x + x_inc
                    p = p + d_y

                y = y + y_inc


    def circle(self, x_c, y_c, radius, color, width):
        x = 0
        y = radius
        p = 5 / (4 - radius)

        self.point(x, y, color, width)

        while x < y:
            x = x + 1

            if p < 0:
                p = p + 2 * x + 1
            else:
                y = y - 1
                p = p + 2 * x + 1 - 2 * y
            
            x = x + x_c
            y = y + y_c

            self.point(x + x_c, y + y_c, color, width)
            self.point(y + x_c, x + y_c, color, width)
            self.point(y + x_c, -x + y_c, color, width)
            self.point(-x + x_c, y + y_c, color, width)
            self.point(-x + x_c, -y + y_c, color, width)
            self.point(-y + x_c, -x + y_c, color, width)
            self.point(-y + x_c, x + y_c, color, width)
            self.point(x + x_c, -y + y_c, color, width)


    def paint(self, x, y, color):
        if "black" == color(x, y): 
            self.point(x, y, color, 2)
            self.paint(x + 1, y, color)
            self.paint(x - 1, y, color)
            self.paint(x, y + 1, color)
            self.paint(x, y - 1, color)


    def projection(self, x, y, z, f = 100):
        F = f * 50
        x_der = (x * f)/(F - z)
        y_der = (y * f)/(F - z)
        return (x_der, y_der)


    def text(self, x, y, word, color=color_rgb(28, 51, 59), width=10, style="bold"):
        x_text = x + self.x_max/ 2
        y_text = self.y_max / 2 - y

        t = Text(Point(x_text, y_text), word)
        t.setOutline(color)
        t.setSize(width)
        t.setStyle(style)
        t.draw(self.win)

    
    def direction(self, x, y, dir):
        if(x == 0):
            return 0
        
        #Caso decole, adicionar 180 graus
        m = y / x
        a = math.atan(m) 

        if(dir == "P" and x > 0):
            a += math.pi
        elif(dir == "D" and x < 0):
            a += math.pi

        return a


    def rotation(self, x_graph, y_graph, x, y, ang):
        x_der = x * math.cos(ang) - y * math.sin(ang)
        y_der = y * math.cos(ang) + x * math.sin(ang)

        return (int(x_der + x_graph), int(y_der + y_graph))


    def airplaneIcon(self, x_graph, y_graph, dir, text):
        if(dir == "D"):
            color = "red"
        elif(dir == "P"):
            color = "white"
        elif(dir == "S"):
            color = "green"

        x, y = 0, 0

        ang = self.direction(x_graph, y_graph, dir)

        x1, y1 = self.rotation(x_graph, y_graph, x, y, ang)
        x2, y2 = self.rotation(x_graph, y_graph, x-20, y, ang)

        x3, y3 = self.rotation(x_graph, y_graph, x-6, y, ang)
        x4, y4 = self.rotation(x_graph, y_graph, x-8, y+8, ang)

        x5, y5 = self.rotation(x_graph, y_graph, x-6, y, ang)
        x6, y6 = self.rotation(x_graph, y_graph, x-8, y-8, ang)

        x7, y7 = self.rotation(x_graph, y_graph, x-16, y, ang)
        x8, y8 = self.rotation(x_graph, y_graph, x-18, y+4, ang)

        x9, y9 = self.rotation(x_graph, y_graph, x-16, y, ang)
        x10, y10 = self.rotation(x_graph, y_graph, x-18, y-4, ang)

        self.straight(x1, y1, x2, y2, color, 2)
        self.straight(x3, y3, x4, y4, color, 2)
        self.straight(x5, y5, x6, y6, color, 2)
        self.straight(x7, y7, x8, y8, color, 2)
        self.straight(x9, y9, x10, y10, color, 2)

        self.text(x_graph + 35, y_graph + 35, text, color)


    def backgroundScreen(self):
        #Retas do Arco
        self.straight(0, -(size / 2) + 100, 0, (size / 2) - 100, color_rgb(28, 51, 59), 2, "dashed")
        self.straight(-(size / 2) + 100, 0, (size / 2) -100, 0, color_rgb(28, 51, 59), 2, "dashed")
        self.straight(-(size / 2) + 220, (size / 2) - 220, (size / 2) - 220, -(size / 2) + 220, color_rgb(28, 51, 59), 2, "dotted")
        self.straight((size / 2) - 220, (size / 2) - 220, -(size / 2) + 220, -(size / 2) + 220, color_rgb(28, 51, 59), 2, "dotted")

        #Círculos do Arco
        self.circle(0, 0, size / 2 - 100, color_rgb(28, 51, 59), 1)
        self.circle(0, 0, size / 2 - 200, color_rgb(28, 51, 59), 1)
        self.circle(0, 0, size / 2 - 300, color_rgb(28, 51, 59), 1)
        self.circle(0, 0, size / 2 - 400, color_rgb(28, 51, 59), 1)

        #Aeroportos
        self.straight(100, 200, 200, 90, color_rgb(28, 51, 59), 2)

        self.backgroundScreenText()


    def backgroundScreenText(self):
        #Graus
        self.text((size / 2) -50, 0, "0°/360°")
        self.text((size / 2) - 170, (size / 2) - 170, "45°")
        self.text(-(size / 2) + 170, (size / 2) - 170, "135°")
        self.text(0, (size / 2) -50, "90°")
        self.text(-(size / 2) + 50, 0, "180°")
        self.text(-(size / 2) + 170, -(size / 2) + 170, "225°")
        self.text(0, -(size / 2) + 50, "270°")
        self.text((size / 2) - 170, -(size / 2) + 170, "315°")

        #Aeroportos
        self.text(150, 145, "Sao Jose Dos Campos", color_rgb(49, 50, 98))


    def saveScreen(self):
        self.background = deepcopy(self.z_buffer)


    def reloadScreen(self):
        self.win.delete("all")
        self.win.update()

        for x in range(self.x_max):
            for y in range(self.y_max):
                if(self.background[int(x)][int(y)] == "empty"):
                    continue

                self.win.plotPixel(x, y, self.background[int(x)][int(y)])

        self.backgroundScreenText()


    def scan(self, starter):
        datas_scan = []
        last_index = 0
        
        for i in range(int(starter), len(self.datas)):
            if(i != int(starter)):

                if(self.datas[i][0] != self.datas[i-1][0]):
                    last_index = i
                    break

            datas_scan.append(self.datas[i])

        return [datas_scan, last_index]


    def readScan(self, datas):
        for data in datas:
            x,y = self.projection(data[5], data[6], data[7])
            self.airplaneIcon(x, y, data[1], data[2])


# ------------------------- RUNNING PROGRAM -------------------------  #
size = 1000
primitive = PrimitiveGraphs(size, size)

primitive.backgroundScreen()
primitive.saveScreen()

scan_datas, last_index = primitive.scan(0)
primitive.readScan(scan_datas)

time.sleep(5)

while True:
    primitive.reloadScreen()

    scan_datas, last_index = primitive.scan(last_index)
    primitive.readScan(scan_datas)

    time.sleep(5)

    if(last_index == 0):
        time.sleep(5)
        primitive.win.close()
        break